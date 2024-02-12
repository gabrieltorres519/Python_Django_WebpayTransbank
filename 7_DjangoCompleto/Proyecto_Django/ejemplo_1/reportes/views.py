from django.shortcuts import render
from home.models import *
from django.http import HttpResponse, Http404, HttpResponseRedirect
from django.contrib import messages
from django.conf import settings
from django.template.loader import render_to_string
from weasyprint import HTML # pip intall WeasyPrint
import os
from os import remove
from .forms import *
from django.core.files.storage import FileSystemStorage
import xlrd # pip install xlrd | pip install django-excel
from openpyxl import load_workbook # pip install openpyxl
import django_excel as excel

# Create your views here.
def reportes_inicio(request):
    # return HttpResponse("Hola mundo")
    return  render(request, 'reportes/home.html', {}) # contiene la ruta del template de esa vista y los datos que se quieren renderizar en la vista

def reportes_exportar_excel(request):
	return render(request, 'reportes/exportar_excel.html', {})

def reportes_exportar_excel_ejecutar(request):
	export = []
	export.append(['ID', 'NOMBRE', 'NÚMERO'])
	datos = Nombres.objects.all()
	for dato in datos:
		export.append([dato.id, dato.nombre, dato.numero])
	sheet = excel.pe.Sheet(export)
	return excel.make_response(sheet, "csv", file_name="ejemplo_de_excel.csv")

def reportes_importar_txt(request):
	if request.method =='POST':
		form = Formulario_importar_txt(request.POST, request.FILES)
		if form.is_valid():
			datos = form.cleaned_data
			#guardamos el archivo
			myfile = request.FILES['file']
			fs=FileSystemStorage()
			filename=fs.save(f"txt/archivo.txt", myfile)
			uploaded_file_url=fs.url(filename)

			fichero = open(settings.BASE_DIR+'/media/txt/archivo.txt')
			lineas = fichero.readlines()
			for linea in lineas:
				Tracking.objects.create(descripcion=linea)
			remove(settings.BASE_DIR+'/media/txt/archivo.txt')
			messages.add_message(request, messages.SUCCESS, f"El txt se importó exitosamente")
			return HttpResponseRedirect(f"/reportes/importar-txt")

	else:
		form=Formulario_importar_txt()
	return render(request, 'reportes/importar-txt.html', {'form':form})

def reportes_importar_excel(request):
    if request.method == 'POST':
        form = Formulario_importar_excel(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = request.FILES['file']
            fs = FileSystemStorage()
            filename = fs.save("excel/archivo.xlsx", uploaded_file)
            upload_file_url = fs.url(filename)

            # Leer el archivo
            workbook = load_workbook(os.path.join(settings.MEDIA_ROOT, "excel/archivo.xlsx"))
            sheet = workbook.active

            # Guardar la información en la base de datos
            for row in sheet.iter_rows(min_row=2, values_only=True):
                Nombres.objects.create(numero=row[0], nombre=row[1])

            os.remove(os.path.join(settings.MEDIA_ROOT, "excel/archivo.xlsx"))

            messages.success(request, "El excel se importó exitosamente")
            return HttpResponseRedirect("/reportes/importar-excel")
    else:
        form = Formulario_importar_excel()

    return render(request, 'reportes/importar_excel.html', {'form': form})

def reportes_pdf(request):
    ruta = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    base_url = settings.BASE_URL
    template = f"{ruta}/templates/layout/ajax.html"
    texto="mi muñeca me habló"
    data = {'texto': texto, 'template':template, 'request': request, 'ruta': ruta, 'base_url': base_url}

    html = render_to_string(f"{ruta}/templates/reportes/pdf.html", data)
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"]= "inline; mi_pdf.pdf"
    HTML(string=html).write_pdf(response)
    return response
